import base64
from datetime import datetime
import os
from shutil import copyfile

import dateutil.parser
from django.apps import apps
from django.conf import settings
from django.core.mail import send_mail
from django.shortcuts import get_object_or_404
from django.template.defaultfilters import slugify
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Color
from openpyxl.styles.fills import PatternFill
from rest_framework import generics
from rest_framework.pagination import PageNumberPagination
from rest_framework.parsers import FileUploadParser, MultiPartParser
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView

from . import models
from . import serializers


class StandardResultsSetPagination(PageNumberPagination):
    page_size = 25
    page_size_query_param = 'page_size'
    max_page_size = 25


class ReportDetail(generics.RetrieveDestroyAPIView):

    serializer_class = serializers.ReportSerializer
    permissions = [IsAuthenticated]

    def get_object(self):
        return get_object_or_404(models.Report, id=self.kwargs.get("report_id"))


class ReportList(generics.ListAPIView):

    serializer_class = serializers.ReportSerializer
    permissions = [IsAuthenticated]
    queryset = models.Report.objects.all()


class ReportPublish(APIView):

    permissions = [IsAuthenticated]

    def put(self, request, *args, **kwargs):
        print(kwargs)
        report = get_object_or_404(models.Report, id=kwargs["report_id"])
        report.draft = False
        report.save()
        return Response(status=200, data=serializers.ReportSerializer(report).data)


class CustomerServiceList(generics.ListCreateAPIView):

    serializer_class = serializers.CustomerServiceSerializer
    permissions = [IsAuthenticated]

    def get_queryset(self):
        return models.CustomerService.objects.filter(
            team=self.request.user.team
        ).filter(
            draft=False
        ).order_by("-created_at")

    def perform_create(self, serializer):
        report = serializer.save(author=self.request.user)
        if not report.draft:
            static_path = os.path.join(settings.BASE_DIR, 'staticfiles/xlsx/csqr/xlxs')
            print(static_path)
            print(static_path.isfile())
            send_mail(f"New CSQR for {report.company_name}", "Hello There",
                      None, ["devin.s@priority1pos.com"])


class CustomerServiceSimpleDraftsList(generics.ListAPIView):

    serializer_class = serializers.CustomerServiceSimpleSerializer
    permissions = [IsAuthenticated]
    pagination_class = StandardResultsSetPagination

    def get_queryset(self):
        return models.CustomerService.objects.filter(
            author=self.request.user
        ).filter(draft=True).order_by("-created_at")


class CustomerServiceRecentDraftsList(generics.ListAPIView):

    serializer_class = serializers.CustomerServiceSimpleSerializer
    permissions = [IsAuthenticated]

    def get_queryset(self):
        return models.CustomerService.objects.filter(
            author=self.request.user
        ).filter(draft=True).order_by("-created_at")[:5]


class CustomerServiceRecentList(generics.ListAPIView):

    serializer_class = serializers.CustomerServiceSimpleSerializer
    permissions = [IsAuthenticated]

    def get_queryset(self):
        return models.CustomerService.objects.filter(
            author=self.request.user
        ).filter(draft=False).order_by("-created_at")[:5]


class CustomerServiceRetrieveUpdate(generics.RetrieveUpdateAPIView):

    serializer_class = serializers.CustomerServiceNestedSerializer
    permissions = [IsAuthenticated]
    queryset = models.CustomerService.objects.all()

    def update(self, request, *args, **kwargs):
        update_response = super().update(request, *args, **kwargs)
        report = update_response.data
        
        if not report["draft"]:
            user_model = apps.get_model("accounts", "Account")
            template_spread = os.path.join(settings.BASE_DIR, 'staticfiles/xlsx/csqr.xlsx')
            report_dir = os.path.join(settings.MEDIA_ROOT, f"reports/csqr/")
            if not os.path.exists(report_dir):
                os.makedirs(report_dir)
            new_spread = os.path.join(report_dir, f"{report['id']}.xlsx")
            copyfile(template_spread, new_spread)
            wb = load_workbook(filename = new_spread)
            ws = wb.active

            ws["D4"] = datetime.now() # Date
            ws["A7"] = report["company_name"]
            ws["C7"] = report["service_type"]
            ws["A8"] = report["location"]
            ws["C8"] = report["description"]
            ws["A9"] = report["client_name"]

            row = 13
            for record in report["time_records"]:
                ws.insert_rows(row)
                ws.merge_cells(f"D{row}:F{row}")

                start = dateutil.parser.parse(record["start"])
                end = dateutil.parser.parse(record["end"])

                ws[f"A{row}"] = start.strftime("%x")
                ws[f"B{row}"] = start.strftime("%X")
                ws[f"C{row}"] = end.strftime("%X")

                employee_string = ""
                for employee in record["users"]:
                    employee_account = user_model.objects.get(id=employee)
                    employee_string += f"{employee_account.first_name} {employee_account.last_name}, "
                ws[f"D{row}"] = employee_string[:len(employee_string)-2]
                
                total_hours = (end - start).total_seconds()/3600
                ws[f"G{row}"] = round(total_hours, 2)

                row += 1

            row += 3
            items = {}
            for item in report["inventory_checkouts"]:
                ws.insert_rows(row)
                ws.merge_cells(f"A{row}:B{row}")
                ws.merge_cells(f"C{row}:D{row}")
                ws.merge_cells(f"E{row}:G{row}")

                ws[f"A{row}"] = item["description"]
                ws[f"C{row}"] = item["model"]
                ws[f"E{row}"] = item["serial"]

                if item["model"] in items:
                    items[item["model"]]["quantity"] += 1
                else:
                    items[item["model"]] = {"description": item["description"], "quantity": 1}

                row += 1

            row += 3
            for model in items:
                ws.insert_rows(row)
                ws.merge_cells(f"A{row}:B{row}")
                ws.merge_cells(f"D{row}:D{row}")
                ws.merge_cells(f"G{row}:G{row}")

                ws[f"A{row}"] = items[model]["description"]
                ws[f"D{row}"] = model
                ws[f"G{row}"] = items[model]["quantity"]

                row += 1
            
            row += 2

            ws[f"A{row}"] = report["summary"]
            ws[f"A{row}"].alignment = Alignment(wrap_text=True)
            ws.merge_cells(f"A{row}:G{row+4}")

            row += 7
            
            selectedColor = Color(rgb="FF0000")
            selectedFill = PatternFill(patternType="solid", fgColor=selectedColor)

            ws[f"A{row}"].fill = selectedFill

            wb.save(new_spread)
            # send_mail(f"New CSQR for {report['company_name']}", "Hello There",
            #           None, ["devin.s@priority1pos.com", "alyssa@priority1pos.com"])
        return update_response


class CustomerServiceSimpleList(generics.ListAPIView):

    serializer_class = serializers.CustomerServiceSimpleSerializer
    permissions = [IsAuthenticated]
    pagination_class = StandardResultsSetPagination

    def get_queryset(self):
        print(self.request.user)
        print(self.request.user.team)
        return models.CustomerService.objects.filter(
            team=self.request.user.team
        ).filter(
            draft=False
        ).order_by("-created_at")


class InventoryCheckOutListCreate(generics.ListCreateAPIView):

    serializer_class = serializers.InventoryCheckOutSerializer
    permissions = [IsAuthenticated]

    def get_queryset(self):
        return models.InventoryCheckOut.objects.filter(report=self.kwargs.get("report_id")).order_by("-created_at")[:5]

    def get_serializer(self, *args, **kwargs):
        if "data" in kwargs:
            data = kwargs["data"]

            # check if many is required
            if isinstance(data, list):
                kwargs["many"] = True

        return super(InventoryCheckOutListCreate, self).get_serializer(*args, **kwargs)

    def create(self, request, *args, **kwargs):
        print(request.data)
        for inventory in request.data:
            inventory["report"] = kwargs["report_id"]
        return super(InventoryCheckOutListCreate, self).create(request)


class InventoryClear(APIView):

    permissions = [IsAuthenticated]

    def delete(self, request, *args, **kwargs):
        models.InventoryCheckOut.objects.filter(
            report=self.kwargs["report_id"]).delete()
        return Response(status=204)


class SignatureRetreiveDelete(generics.RetrieveUpdateDestroyAPIView):

    serializer_class = serializers.SignatureSerializer
    permissions = [IsAuthenticated]

    def get_object(self):
        return get_object_or_404(models.Signature, id=self.kwargs["signature_id"])


class SignatureCreate(generics.CreateAPIView):

    serializer_class = serializers.SignatureSerializer
    permissions = [IsAuthenticated]


class TimeEntryListCreate(generics.ListCreateAPIView):

    serializer_class = serializers.TimeEntrySerializer
    permissions = [IsAuthenticated]

    def get_queryset(self):
        return models.TimeEntry.objects.filter(report=self.kwargs.get("report_id"))

    def get_serializer(self, *args, **kwargs):
        if "data" in kwargs:
            data = kwargs["data"]

            # check if many is required
            if isinstance(data, list):
                kwargs["many"] = True

        return super(TimeEntryListCreate, self).get_serializer(*args, **kwargs)

    def create(self, request, *args, **kwargs):
        print(request.data)
        if isinstance(request.data, list):
            for inventory in request.data:
                inventory["report"] = kwargs["report_id"]
        else:
            request.data["report"] = kwargs["report_id"]
        return super(TimeEntryListCreate, self).create(request)


class TimeEntryClear(APIView):

    permissions = [IsAuthenticated]

    def delete(self, request, *args, **kwargs):
        models.TimeEntry.objects.filter(
            report=self.kwargs["report_id"]).delete()
        return Response(status=204)
