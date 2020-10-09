from datetime import datetime
import os
from shutil import copyfile

from django.apps import apps
from django.conf import settings
from django.core.mail import EmailMessage
from django.utils.text import slugify
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment, Color, Font
from openpyxl.styles.fills import PatternFill
from PIL import Image as PILImage
from . import models


def setup_file_structure(id):
    template_spread = os.path.join(
        settings.BASE_DIR, 'staticfiles/xlsx/csqr.xlsx'
    )
    report_dir = os.path.join(settings.MEDIA_ROOT, f"reports/csqr/")

    if not os.path.exists(report_dir):
        os.makedirs(report_dir)

    new_spread = os.path.join(report_dir, f"{id}.xlsx")
    copyfile(template_spread, new_spread)
    return new_spread


def write_additional_info(report, ws, row):
    selectedColor = Color(rgb="FF169f43")
    unselectedColor = Color(rgb="FFF7fAFC")
    whiteFont = Font(color="FFFFFFFF")
    selectedFill = PatternFill(
        patternType="solid", fgColor=selectedColor
    )
    unselectedFill = PatternFill(
        patternType="solid", fgColor=unselectedColor
    )

    if report.billable:
        ws[f"A{row}"].fill = selectedFill
        ws[f"A{row}"].font = whiteFont
    else:
        ws[f"A{row}"].fill = unselectedFill
    ws.merge_cells(f"A{row}:C{row}")

    if report.completed:
        ws[f"E{row}"].fill = selectedFill
        ws[f"E{row}"].font = whiteFont
    else:
        ws[f"E{row}"].fill = unselectedFill
    ws.merge_cells(f"E{row}:G{row}")

    row += 1

    if report.tested:
        ws[f"A{row}"].fill = selectedFill
        ws[f"A{row}"].font = whiteFont
    else:
        ws[f"A{row}"].fill = unselectedFill
    ws.merge_cells(f"A{row}:C{row}")

    if report.pictures:
        ws[f"E{row}"].fill = selectedFill
        ws[f"E{row}"].font = whiteFont
    else:
        ws[f"E{row}"].fill = unselectedFill
    ws.merge_cells(f"E{row}:G{row}")

    row += 1

    if report.reviewed:
        ws[f"A{row}"].fill = selectedFill
        ws[f"A{row}"].font = whiteFont
    else:
        ws[f"A{row}"].fill = unselectedFill
    ws.merge_cells(f"A{row}:C{row}")

    if report.satisfied:
        ws[f"E{row}"].fill = selectedFill
        ws[f"E{row}"].font = whiteFont
    else:
        ws[f"E{row}"].fill = unselectedFill
    ws.merge_cells(f"E{row}:G{row}")

    row += 1

    ws[f"C{row}"] = report.followup
    ws[f"C{row}"].alignment = Alignment(horizontal="left")
    ws.merge_cells(f"C{row}:G{row}")

    return row


def write_signature(report, ws, row):
    blob = settings.BUCKET.blob(report.signature.ref)
    signature_dir = os.path.join(
        settings.MEDIA_ROOT, f"reports/signatures/"
    )
    if not os.path.exists(signature_dir):
        os.makedirs(signature_dir)
    signature_file = os.path.join(
        signature_dir, f"{report.signature.id}.png"
    )
    blob.download_to_filename(os.path.join(signature_file))
    signature_img = PILImage.open(signature_file)
    size_ratio = signature_img.height / signature_img.width
    new_height = 50
    new_width = int(new_height / size_ratio)
    new_signature_img = signature_img.resize(
        (new_width, new_height), PILImage.BICUBIC)
    new_signature_img.save(signature_file)

    img = Image(signature_file)
    ws[f"A{row}"] = ""
    ws.add_image(img, f"A{row}")
    ws.merge_cells(f"A{row}:G{row+2}")

    row += 3

    ws[f"A{row}"] = datetime.now().strftime("%a %B %-d, %Y %-I:%M %p")

    return row


def write_spread_header(report, ws):
    ws["D4"] = datetime.now()  # Date
    ws["A7"] = report.company_name
    ws["C7"] = report.service_type
    ws["A8"] = report.location
    ws["C8"] = report.description
    ws["A9"] = report.client_name


def write_time_records(report, ws, row):
    for record in report.time_records.all():
        ws.insert_rows(row)
        ws.merge_cells(f"D{row}:F{row}")

        ws[f"A{row}"] = record.start.strftime("%x")
        ws[f"B{row}"] = record.start.strftime("%X")
        ws[f"C{row}"] = record.end.strftime("%X")

        employee_string = ""

        for employee in record.users.all():
            employee_string += f"{employee.first_name} {employee.last_name}, "
        ws[f"D{row}"] = employee_string[:len(employee_string)-2]

        total_hours = (record.end - record.start).total_seconds()/3600
        ws[f"G{row}"] = round(total_hours, 2)

        row += 1
    return row


def write_inventory(report, ws, row):
    items = {}
    for item in report.inventory_checkouts.all():
        ws.insert_rows(row)
        ws.merge_cells(f"A{row}:B{row}")
        ws.merge_cells(f"C{row}:D{row}")
        ws.merge_cells(f"E{row}:G{row}")

        ws[f"A{row}"] = item.description
        ws[f"C{row}"] = item.model
        ws[f"E{row}"] = item.serial

        if item.model in items:
            items[item.model]["quantity"] += 1
        else:
            items[item.model] = {
                "description": item.description, "quantity": 1}

        row += 1

    row += 3
    for model in items:
        ws.insert_rows(row)
        ws.merge_cells(f"A{row}:B{row}")
        ws.merge_cells(f"D{row}:D{row}")
        ws.merge_cells(f"G{row}:G{row}")

        ws[f"A{row}"] = items[model]["description"]
        ws[f"D{row}"] = model
        ws[f"G{row}"] = items[model]["quantity"]

        row += 1

    return row


def build_spread(report_id):
    report = models.CustomerService.objects.get(id=report_id)
    user_model = apps.get_model("accounts", "Account")
    spread_file = setup_file_structure(report.id)

    wb = load_workbook(filename=spread_file)
    ws = wb.active

    write_spread_header(report, ws)

    row = 13

    row = write_time_records(report, ws, row)

    row += 3

    row = write_inventory(report, ws, row)

    row += 2

    ws[f"A{row}"] = report.summary
    ws[f"A{row}"].alignment = Alignment(
        horizontal="left", wrap_text=True, vertical="top"
    )
    ws.merge_cells(f"A{row}:G{row+4}")

    row += 7

    row = write_additional_info(report, ws, row)

    row += 4

    ws[f"A{row}"] = report.client_name

    row += 1

    print(report.signature)
    if report.signature:
        row = write_signature(report, ws, row)

    wb.save(spread_file)

    upload_spread(report.id, report.team.slug, spread_file)

    return {"report": report, "spread_file": spread_file}


def email_spread(report, spread_file, additional_recipients=[], updated_model=False):
    user_model = apps.get_model("accounts", "Account")
    report_admins = user_model.objects.filter(
        team=report.team
    ).filter(report_admin=True)

    if report.billable:
        billable_icon = "\u2713"
    else:
        billable_icon = "\u2717"

    if report.completed:
        completed_icon = "\u2713"
    else:
        completed_icon = "\u2717"

    if updated_model:
        message = f"""
A report has been updated by a minion admin for {report.company_name}.

{report.company_name}
{report.location}
{report.client_name}
{report.service_type}
{report.description}

{report.summary}

Billable: {billable_icon}
Completed: {completed_icon}

Followup: {report.followup} 

ID: {report.id}
        """
        subject = f"CSQR for {report.company_name} has been updated."
    else:
        message = f"""
A new report for {report.company_name} has been generated by Minion.

{report.company_name}
{report.location}
{report.client_name}
{report.service_type}
{report.description}

{report.summary}

Billable: {billable_icon}
Completed: {completed_icon}

Followup: {report.followup} 

ID: {report.id}
        """
        subject = f"New CSQR for {report.company_name}"

    email_list = []
    for user in report_admins:
        email_list.append(user.email)
    for email in additional_recipients:
        email_list.append(email)
    email_list.append(report.author.email)
    email_list.append(report.last_edited_by.email)
    email_list = list(dict.fromkeys(email_list))
    print(email_list)
    admin_message = EmailMessage(
        subject,
        message,
        settings.EMAIL_HOST_USER,
        email_list,
    )
    with open(spread_file, mode="rb") as spread_data:
        admin_message.attach(
            f"{slugify(report.company_name)}_{slugify(report.client_name)}_{report.id}.xlsx", spread_data.read()
        )
    admin_message.send()


def upload_spread(report_id, team_slug, spread_file):
    blob = settings.BUCKET.blob(
        f"{team_slug}/reports/csqr/{report_id}.xlsx"
    )
    blob.upload_from_filename(spread_file)
