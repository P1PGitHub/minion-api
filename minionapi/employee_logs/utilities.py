from datetime import datetime, timedelta
import os
import pytz
from shutil import copyfile

from django.apps import apps
from django.conf import settings
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment, Border, Font, NamedStyle, Side
from PIL import Image as PILImage

from . import models


DATE_HEADER_STYLE = NamedStyle(name="Date Header")
STANDARD_STYLE = NamedStyle(name="Standard")
BOTTOM_BORDER = Border(bottom=Side(border_style="thin", color="FF000000"))
BOLD_FONT = Font(bold=True)
RESOLVED_FONT = Font(bold=True, color="FF276749")
UNRESOLVED_FONT = Font(bold=True, color="FF9C4221")
RIGHT_ALIGN = Alignment(horizontal="right", wrap_text=True, vertical="top")
DATE_HEADER_STYLE.border = BOTTOM_BORDER
DATE_HEADER_STYLE.font = BOLD_FONT
STANDARD_STYLE.alignment = Alignment(
    horizontal="left", wrap_text=True, vertical="top"
)


def setup_file_structure(start, end, user):
    template_spread = os.path.join(
        settings.BASE_DIR, 'staticfiles/xlsx/activity.xlsx'
    )
    spread_dir = os.path.join(
        settings.MEDIA_ROOT, f"activity/{user.team.slug}/")

    if not os.path.exists(spread_dir):
        os.makedirs(spread_dir)

    new_spread = os.path.join(
        spread_dir, f"{start.strftime('%Y%M%d')}-{end.strftime('%Y%M%d')}-{user.last_name}-{user.first_name}.xlsx")
    copyfile(template_spread, new_spread)
    return new_spread


def setup_team_file_structure(start, end, team):
    template_spread = os.path.join(
        settings.BASE_DIR, 'staticfiles/xlsx/activity.xlsx'
    )
    spread_dir = os.path.join(
        settings.MEDIA_ROOT, f"activity/{team.slug}/"
    )

    if not os.path.exists(spread_dir):
        os.makedirs(spread_dir)

    new_spread = os.path.join(
        spread_dir, f"{start.strftime('%Y%M%d')}-{end.strftime('%Y%M%d')}.xlsx")
    copyfile(template_spread, new_spread)
    return new_spread


def write_date_header(ws, start, end):
    if end.strftime("%s") == start.strftime("%s"):
        ws["D4"] = start.strftime("%A %b %-d, %Y")
    else:
        ws["D4"] = start.strftime("%A %b %-d, %Y") + \
            " - " + end.strftime("%A %b %-d, %Y")


def write_employee_information(ws, user):
    ws["A7"] = f"{user.last_name}, {user.first_name}"
    ws["A8"] = user.team.name


def write_logo(ws, team):
    logo_path = os.path.join(
        settings.MEDIA_ROOT, f"temp/logo/{team.slug}/{team.logo_ref}"
    )
    if not os.path.exists(logo_path):
        os.makedirs(
            os.path.join(
                settings.MEDIA_ROOT, f"temp/logo/{team.slug}/"
            )
        )
        blob = settings.BUCKET.blob(f"{team.slug}/logo/{team.logo_ref}")
        blob.download_to_filename(logo_path)
        logo_image = PILImage.open(logo_path)
        size_ratio = logo_image.height / logo_image.width
        new_height = 75
        new_width = int(new_height / size_ratio)
        new_logo_image = logo_image.resize(
            (new_width, new_height), PILImage.BICUBIC)
        new_logo_image.save(logo_path)
        ws_img = Image(logo_path)
        ws["A1"] = ""
        ws.add_image(ws_img, "A1")
        ws.merge_cells("A1:C4")
    else:
        ws_img = Image(logo_path)
        ws["A1"] = ""
        ws.add_image(ws_img, "A1")
        ws.merge_cells("A1:C4")


def write_time_entries(ws, time_entries):
    row = 12
    companies = {}
    day_total_hours = 0
    employee_total_hours = 0
    for index, entry in enumerate(time_entries):
        ws.insert_rows(row)
        entry_start_tz = entry.start.astimezone(
            pytz.timezone(entry.team.timezone))
        entry_end_tz = entry.end.astimezone(pytz.timezone(entry.team.timezone))
        if index > 0 and not time_entries[index-1].start.astimezone(pytz.timezone(entry.team.timezone)).strftime("%x") == entry_start_tz.strftime("%x"):
            ws[f"G{row}"] = round(day_total_hours, 2)
            ws[f"G{row-1}"].border = Border(
                bottom=Side(style="double", color="FF000000")
            )
            day_total_hours = 0
            row += 1
            ws.insert_rows(row)
        if index == 0 or not time_entries[index-1].start.astimezone(pytz.timezone(entry.team.timezone)).strftime("%x") == entry_start_tz.strftime("%x"):
            ws.merge_cells(f"A{row}:G{row}")
            ws[f"A{row}"] = entry_start_tz.strftime("%A %b %-d, %Y")
            for cell in ws[f"A{row}:G{row}"][0]:
                cell.style = DATE_HEADER_STYLE
            ws[f"A{row}"].style = DATE_HEADER_STYLE
            row += 1
            ws.insert_rows(row)

        ws[f"A{row}"] = entry.client_name
        ws[f"B{row}"] = entry.company_name
        ws[f"C{row}"] = entry_start_tz.strftime("%X")
        ws[f"D{row}"] = entry_end_tz.strftime("%X")
        ws[f"E{row}"] = entry.description
        ws[f"F{row}"] = entry.summary
        total_hours = (entry_end_tz - entry_start_tz).total_seconds()/3600
        ws[f"G{row}"] = round(total_hours, 2)
        for cell in ws[f"A{row}:G{row}"][0]:
            cell.style = STANDARD_STYLE
        ws[f"G{row}"].alignment = RIGHT_ALIGN
        if entry.resolved:
            ws[f"A{row}"].font = RESOLVED_FONT
            ws[f"B{row}"].font = RESOLVED_FONT
        else:
            ws[f"A{row}"].font = UNRESOLVED_FONT
            ws[f"B{row}"].font = UNRESOLVED_FONT

        if entry.company_name in companies:
            companies[entry.company_name]["quantity"] += 1
            companies[entry.company_name]["total_hours"] += total_hours
            if entry.resolved:
                companies[entry.company_name]["resolved"] += 1
            else:
                companies[entry.company_name]["unresolved"] += 1
        else:
            companies[entry.company_name] = {
                "quantity": 1,
                "resolved": 0,
                "unresolved": 0,
                "total_hours": total_hours
            }
            if entry.resolved:
                companies[entry.company_name]["resolved"] += 1
            else:
                companies[entry.company_name]["unresolved"] += 1

        day_total_hours += total_hours
        employee_total_hours += total_hours
        row += 1

    ws[f"G{row}"] = round(day_total_hours, 2)
    ws[f"G{row-1}"].border = Border(
        bottom=Side(style="double", color="FF000000")
    )
    ws.insert_rows(row+1)

    row += 4
    quantities = {
        "total": 0,
        "resolved": 0,
        "unresolved": 0
    }
    for name in sorted(companies.keys()):
        ws.merge_cells(f"A{row}:C{row}")
        ws[f"A{row}"] = name
        ws[f"D{row}"] = companies[name]["quantity"]
        ws[f"E{row}"] = companies[name]["resolved"]
        ws[f"F{row}"] = companies[name]["unresolved"]
        ws[f"G{row}"] = round(companies[name]["total_hours"], 2)
        quantities["total"] += companies[name]["quantity"]
        quantities["resolved"] += companies[name]["resolved"]
        quantities["unresolved"] += companies[name]["unresolved"]
        row += 1

    for cell in ws[f"D{row-1}:G{row-1}"][0]:
        cell.border = Border(
            bottom=Side(style="double", color="FF000000")
        )
    ws[f"D{row}"] = quantities["total"]
    ws[f"E{row}"] = quantities["resolved"]
    ws[f"F{row}"] = quantities["unresolved"]
    ws[f"G{row}"] = round(employee_total_hours, 2)

    return quantities


def write_totals_header(ws, user, quantities):
    now_tz = datetime.now(
        tz=pytz.timezone("UTC")
    ).astimezone(
        pytz.timezone(user.team.timezone)
    )

    ws["F7"] = now_tz.strftime("%x %X %Z")
    ws["G8"] = round(quantities["total"], 2)


def build_sheet(spread_file, user, start, end):
    wb = load_workbook(spread_file)
    template_ws = wb["Sheet1"]
    ws = wb.copy_worksheet(template_ws)
    ws.title = f"{user.last_name}, {user.first_name}"
    write_date_header(ws, start, end)
    write_employee_information(ws, user)
    time_entries = models.WorkEntry.objects.filter(
        start__gte=start).filter(end__lte=end + timedelta(days=1)).filter(user=user).order_by("start")
    quantities = write_time_entries(ws, time_entries)
    write_totals_header(ws, user, quantities)
    write_logo(ws, user.team)
    wb.save(spread_file)
    return wb


def build_user_activity_spread(start, end, user):
    try:
        spread_file = setup_file_structure(start, end, user)
        wb = build_sheet(spread_file, user, start, end)
        wb.remove_sheet(wb["Sheet1"])
        wb.save(spread_file)
    except Exception as e:
        print(e)
    return upload_spread(user, start, end, spread_file)


def build_team_activity_spread(start, end, team):
    try:
        user_model = apps.get_model("accounts", "account")
        members = user_model.objects.filter(
            team=team).order_by("last_name", "first_name")
        spread_file = setup_team_file_structure(start, end, team)
        for member in members:
            wb = build_sheet(spread_file, member, start, end)
        wb.remove_sheet(wb["Sheet1"])
        wb.save(spread_file)
    except Exception as e:
        print(e)
    return upload_team_spread(team, start, end, spread_file)


def upload_spread(user, start, end, spread_file):
    start_string = start.astimezone(pytz.timezone(
        user.team.timezone)).strftime("%Y%m%d")
    end_string = end.astimezone(pytz.timezone(
        user.team.timezone)).strftime("%Y%m%d")
    ref = f"{user.team.slug}/temp/activity/{start_string}-{end_string}-{user.last_name}-{user.first_name}.xlsx"

    blob = settings.BUCKET.blob(ref)
    blob.upload_from_filename(spread_file)
    return {
        "name": f"{start_string}-{end_string}-{user.last_name}-{user.first_name}.xlsx",
        "ref": ref,
        "url": blob.public_url
    }


def upload_team_spread(team, start, end, spread_file):
    start_string = start.astimezone(pytz.timezone(
        team.timezone)).strftime("%Y%m%d")
    end_string = end.astimezone(pytz.timezone(
        team.timezone)).strftime("%Y%m%d")
    ref = f"{team.slug}/temp/activity/{start_string}-{end_string}.xlsx"

    blob = settings.BUCKET.blob(ref)
    blob.upload_from_filename(spread_file)
    return {
        "name": f"{start_string}-{end_string}.xlsx",
        "ref": ref,
        "url": blob.public_url
    }
